"""
Universe-import vanaf de complete noteringslijsten van de beurzen zelf (fase 4).

Bronnen (elk één --source):
  euronext       live.euronext.com CSV-endpoint (auto-download of --file).
                 Amsterdam/Brussel/Parijs/Lissabon/Dublin/Oslo, incl. Growth/Access.
  xetra          Deutsche Börse "Listed companies" xlsx (auto-download of --file).
                 Tabbladen Prime Standard + General Standard (+ Scale met --include-growth).
  nasdaq-nordic  api.nasdaq.com/api/nordic screener, Main Market per beurs
                 (CPH/STO/HEL/ICE). First North wordt bewust NIET opgehaald (zie --help).
  baltic         nasdaqbaltic.com share-list xlsx (auto-download of --file).
                 Baltic Main + Secondary List (+ First North Baltic met --include-growth).
  gpw            Warschau. Officiële gpw.pl is vanaf sommige netwerken geblokkeerd;
                 auto-download gebruikt daarom biznesradar.pl (volledige hoofdmarkt-
                 lijst). Geef bij voorkeur --file met de officiële GPW-export.

Werkwijze:
  - parse bron -> records {ticker, name, isin, market, currency, sector, segment}
  - filter: geldige Yahoo-ticker, bekend suffix, default alleen gereguleerde markt
    (--include-growth neemt ook Growth/Scale/First North Baltic/Access mee)
  - dedupe: binnen de batch op ISIN (voorkeur voor de notering in het thuisland
    van de ISIN), daarna tegen ALLE bestaande tickers in de DB (ook inactieve en
    presumed-delisted — die komen er niet opnieuw in) en tegen bekende
    secondary-listings uit engine/remap_rules.py
  - --limit N (default 250): max N nieuwe tickers per run, conform het
    onboarding-tempo van fase 4 (Yahoo-limieten). De fundamentals-rotatie pakt
    nieuwe tickers automatisch met voorrang op (NULL fetched_date eerst).

Market-cap-filter: de noteringslijsten bevatten géén marktkapitalisatie, dus
een --min-mcap filter bij import is niet mogelijk. Filtering op grootte gebeurt
ná de eerste fetch (market_data.market_cap is dan gevuld); zie fase 4-notities.

Gebruik:
  python import_tickers.py --source euronext --dry-run
  python import_tickers.py --source xetra --limit 250
  DATABASE_URL="postgresql://..." python import_tickers.py --source gpw --file gpw.xlsx
  python import_tickers.py --source nasdaq-nordic --dry-run --out preview.csv

Zonder DATABASE_URL werkt alleen --dry-run (bestaande tickers komen dan uit de
live /api/dashboard); echt importeren vereist de database.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
from datetime import date
from pathlib import Path

import requests

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
APP_URL = "https://stockscreen-janco.fly.dev"

# Yahoo-suffix -> (valuta, market-code zoals in de stocks-tabel)
SUFFIX_INFO = {
    "AS": ("EUR", "NL"), "BR": ("EUR", "BE"), "PA": ("EUR", "FR"),
    "LS": ("EUR", "PT"), "IR": ("EUR", "IE"), "OL": ("NOK", "NO"),
    "DE": ("EUR", "DE"), "ST": ("SEK", "SE"), "CO": ("DKK", "DK"),
    "HE": ("EUR", "FI"), "IC": ("ISK", "IS"), "WA": ("PLN", "PL"),
    "TL": ("EUR", "EE"), "RG": ("EUR", "LV"), "VS": ("EUR", "LT"),
}

# ISIN-landprefix -> Yahoo-suffix van de thuismarkt. Bij dezelfde ISIN op
# meerdere beurzen in één batch wint de notering in het thuisland (daar heeft
# yfinance vrijwel altijd de beste dekking — zie remap_rules voor de bewijslast).
ISIN_HOME_SUFFIX = {
    "NL": "AS", "BE": "BR", "FR": "PA", "PT": "LS", "IE": "IR", "NO": "OL",
    "DE": "DE", "SE": "ST", "DK": "CO", "FI": "HE", "IS": "IC", "PL": "WA",
    "EE": "TL", "LV": "RG", "LT": "VS",
}

# Auto-downloads landen hier; `data/` staat in .gitignore.
DOWNLOAD_DIR = Path(__file__).parent / "data" / "exchange_lists"

TICKER_RE = re.compile(r"^[A-Z0-9\-]{1,12}\.[A-Z]{1,3}$")

# De Euronext-lijst bevat naast gewone aandelen ook losse warrants en
# inschrijvingsrechten ("AB SCIENCE BSA"). Geen aandelen -> overslaan.
NON_EQUITY_NAME_RE = re.compile(r"\b(BSA|WARRANT|WTS?|RTS|RIGHTS?|OBLIG)\b", re.IGNORECASE)

# Let op: een onbekende MIC in de lijst (bv. MERK, Euronext Growth Oslo) maakt
# het endpoint kapot — je krijgt dan een 301 naar een lege pagina. Growth Oslo
# ontbreekt daardoor; dat is alleen relevant met --include-growth.
EURONEXT_URL = (
    "https://live.euronext.com/pd_es/data/stocks/download"
    "?mics=XAMS,XBRU,XPAR,XLIS,XOSL,XOAS,ALXB,ALXP,ALXL,XMLI,TNLA,"
    "ENXB,ENXL,XESM,XMSM,XATL"
    "&display_datapoints=dp_stocks&display_filters=df_stocks"
)
XETRA_PAGE = "https://www.deutsche-boerse-cash-market.com/dbcm-en/instruments-statistics/statistics/listed-companies"
XETRA_PAGE_NEW = "https://www.cashmarket.deutsche-boerse.com/cash-en/Data-Tech/statistics/listed-companies"
NASDAQ_API = "https://api.nasdaq.com/api/nordic/screener/shares"
BALTIC_URL = "https://nasdaqbaltic.com/statistics/en/shares?download=1"
GPW_FALLBACK_URL = "https://www.biznesradar.pl/gielda/akcje_gpw"


def _http_get(url: str, **kw) -> requests.Response:
    # IMPORT_TLS_VERIFY=0 voor netwerken met TLS-interceptie (zelfde reden als
    # `curl -k` in de bestaande scripts).
    verify = os.environ.get("IMPORT_TLS_VERIFY", "1") != "0"
    if not verify:
        import urllib3
        urllib3.disable_warnings()
    r = requests.get(url, headers=UA, timeout=60, verify=verify, **kw)
    r.raise_for_status()
    return r


def make_record(ticker: str, name: str | None, isin: str | None,
                sector: str | None = None, segment: str = "regulated") -> dict | None:
    """Bouwt een genormaliseerd record; None als de ticker ongeldig is."""
    ticker = ticker.strip().upper()
    if not TICKER_RE.match(ticker):
        return None
    suffix = ticker.rsplit(".", 1)[1]
    info = SUFFIX_INFO.get(suffix)
    if not info:
        return None
    currency, market = info
    return {
        "ticker": ticker,
        "name": (name or "").strip() or None,
        "isin": (isin or "").strip().upper() or None,
        "sector": (sector or "").strip() or None,
        "market": market,
        "currency": currency,
        "segment": segment,
    }


# ---------------------------------------------------------------- bronnen

def parse_euronext(text: str) -> list[dict]:
    """CSV (;-gescheiden): Name;ISIN;Symbol;Market;Currency;..."""
    market_to_suffix = {
        "amsterdam": "AS", "brussels": "BR", "paris": "PA", "lisbon": "LS",
        "dublin": "IR", "oslo": "OL",
    }
    records = []
    reader = csv.reader(io.StringIO(text), delimiter=";")
    for row in reader:
        if len(row) < 5:
            continue
        name, isin, symbol, market_label = row[0], row[1], row[2], row[3]
        if not (isin and len(isin) == 12 and isin[:2].isalpha() and symbol):
            continue  # header-/metadataregels
        label = market_label.split(",")[0].strip().lower()
        # "Oslo Børs" / "Euronext Expand Oslo" / "Euronext Growth Oslo" -> oslo
        suffix = next((s for key, s in market_to_suffix.items() if key in label), None)
        if label.startswith("oslo b"):
            suffix = "OL"
        if not suffix:
            continue
        if NON_EQUITY_NAME_RE.search(name or ""):
            continue
        segment = "growth" if ("growth" in label or "access" in label or "expand" in label) else "regulated"
        rec = make_record(f"{symbol}.{suffix}", name, isin, segment=segment)
        if rec:
            records.append(rec)
    return records


def fetch_euronext(file: Path | None) -> list[dict]:
    text = file.read_text(encoding="utf-8-sig") if file else _http_get(EURONEXT_URL).content.decode("utf-8-sig")
    return parse_euronext(text)


def parse_xetra(path: Path, include_growth: bool) -> list[dict]:
    """Tabbladen Prime Standard / General Standard (/ Scale). Kolomnamen in een
    headerrij: ISIN, Trading Symbol, Company, Sector, ..."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tabs = {"Prime Standard": "regulated", "General Standard": "regulated"}
    if include_growth:
        tabs["Scale"] = "growth"
    records = []
    for tab, segment in tabs.items():
        if tab not in wb.sheetnames:
            print(f"  [WAARSCHUWING] tabblad '{tab}' niet gevonden", file=sys.stderr)
            continue
        header_seen = False
        for row in wb[tab].iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not header_seen:
                if cells[:2] == ["ISIN", "Trading Symbol"]:
                    header_seen = True
                continue
            isin, symbol, company = cells[0], cells[1], cells[2]
            sector = cells[3] if len(cells) > 3 else None
            if not (len(isin) == 12 and symbol):
                continue
            rec = make_record(f"{symbol}.DE", company, isin, sector=sector, segment=segment)
            if rec:
                records.append(rec)
    return records


def fetch_xetra(file: Path | None, include_growth: bool) -> list[dict]:
    if not file:
        # De xlsx hangt achter een wisselende blob-URL; zoek hem op de statistiekpagina.
        for page_url in (XETRA_PAGE, XETRA_PAGE_NEW):
            try:
                html = _http_get(page_url).text
            except requests.RequestException:
                continue
            m = re.search(r'href="([^"]*?/resource/blob/[^"]*Listed-companies\.xlsx)"', html)
            if m:
                url = m.group(1)
                if url.startswith("/"):
                    url = "https://www.cashmarket.deutsche-boerse.com" + url
                DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
                file = DOWNLOAD_DIR / "xetra_listed_companies.xlsx"
                file.write_bytes(_http_get(url).content)
                break
        if not file:
            raise SystemExit("Xetra-xlsx niet gevonden op de statistiekpagina; download handmatig en geef --file.")
    return parse_xetra(file, include_growth)


def fetch_nasdaq_nordic(file: Path | None) -> list[dict]:
    """Main Market per beurs via de officiële API. --file wordt genegeerd
    (JSON per markt is niet zinvol handmatig aan te leveren).

    First North bewust overgeslagen: de API levert daarvoor geen beurs per rij
    (geen micCode), dus het Yahoo-suffix is niet betrouwbaar af te leiden.
    """
    suffix_by_market = {"CPH": "CO", "STO": "ST", "HEL": "HE", "ICE": "IC"}
    records = []
    for mkt, suffix in suffix_by_market.items():
        r = _http_get(NASDAQ_API, params={"category": "MAIN_MARKET", "market": mkt})
        r.encoding = "utf-8"
        data = json.loads(r.text)["data"]
        rows = data["instrumentListing"]["rows"]
        total = data["pagination"]["total"]
        if len(rows) < total:
            print(f"  [WAARSCHUWING] {mkt}: {len(rows)}/{total} rijen ontvangen (paginering?)", file=sys.stderr)
        for row in rows:
            symbol = (row.get("symbol") or "").replace(" ", "-")
            rec = make_record(f"{symbol}.{suffix}", row.get("fullName"),
                              row.get("isin"), sector=row.get("sector"))
            if rec:
                records.append(rec)
    return records


def parse_baltic(path: Path, include_growth: bool) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    suffix_by_place = {"TLN": "TL", "RIG": "RG", "VLN": "VS"}
    records = []
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if len(cells) < 6:
            continue
        ticker, name, isin, _ccy, place, segment_label = cells[:6]
        suffix = suffix_by_place.get(place)
        if not suffix:
            continue
        is_fn = "first north" in segment_label.lower()
        if is_fn and not include_growth:
            continue
        rec = make_record(f"{ticker}.{suffix}", name, isin,
                          segment="growth" if is_fn else "regulated")
        if rec:
            records.append(rec)
    return records


def fetch_baltic(file: Path | None, include_growth: bool) -> list[dict]:
    if not file:
        DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
        file = DOWNLOAD_DIR / "baltic_shares.xlsx"
        file.write_bytes(_http_get(BALTIC_URL).content)
    return parse_baltic(file, include_growth)


def fetch_gpw(file: Path | None) -> list[dict]:
    """Officiële GPW-export via --file (xlsx met kolommen Ticker/Name/ISIN of
    vergelijkbaar); anders fallback: hoofdmarktlijst van biznesradar.pl."""
    html = None
    if file and file.suffix.lower() in (".html", ".htm"):
        html = file.read_text(encoding="utf-8", errors="replace")
        file = None
    elif file:
        import openpyxl
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        records = []
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if len(cells) < 2 or not cells[0]:
                continue
            isin = next((c for c in cells if len(c) == 12 and c[:2].isalpha() and c[2:].isalnum()), None)
            rec = make_record(f"{cells[0]}.WA", cells[1], isin)
            if rec:
                records.append(rec)
        return records
    if html is None:
        html = _http_get(GPW_FALLBACK_URL).text
    return fetch_gpw_from_html(html)


def fetch_gpw_from_html(html: str) -> list[dict]:
    """Rijen: <a href="/notowania/SLUG">ABBR (VOLNAAM)</a> — ABBR is de GPW-code
    die Yahoo als ticker gebruikt (11B.WA, CDR.WA, ...). Staat de volledige naam
    gelijk aan de code, dan ontbreekt het deel tussen haakjes ("PZU"); die vorm
    moet mee, anders missen we juist de grootste Poolse namen."""
    seen = {}
    for slug, label in re.findall(r'href="/notowania/([A-Z0-9\-\.]+)"[^>]*>([^<]+)</a>', html):
        label = label.strip()
        m = re.match(r"^([A-Z0-9]{2,6}) \(([^)]+)\)$", label)
        if m:
            abbr, name = m.groups()
        elif re.match(r"^[A-Z0-9]{2,6}$", label):
            abbr, name = label, slug.replace("-", " ").title()
        else:
            continue
        rec = make_record(f"{abbr}.WA", name, None)
        if rec:
            seen[rec["ticker"]] = rec
    return list(seen.values())


# ---------------------------------------------------------------- dedupe

def dedupe_isin(records: list[dict]) -> tuple[list[dict], list[str]]:
    """Zelfde ISIN meermaals in de batch -> houd de thuisland-notering."""
    by_isin: dict[str, dict] = {}
    no_isin: list[dict] = []
    dropped: list[str] = []
    for rec in records:
        isin = rec["isin"]
        if not isin:
            no_isin.append(rec)
            continue
        prev = by_isin.get(isin)
        if prev is None:
            by_isin[isin] = rec
            continue
        home = ISIN_HOME_SUFFIX.get(isin[:2])
        cur_suffix = rec["ticker"].rsplit(".", 1)[1]
        prev_suffix = prev["ticker"].rsplit(".", 1)[1]
        if cur_suffix == home and prev_suffix != home:
            dropped.append(prev["ticker"])
            by_isin[isin] = rec
        else:
            dropped.append(rec["ticker"])
    # dubbele tickers over bronnen heen kunnen niet (één run = één bron), maar
    # dezelfde ticker kan wel 2x in een bronbestand staan
    out: dict[str, dict] = {}
    for rec in list(by_isin.values()) + no_isin:
        out.setdefault(rec["ticker"], rec)
    return list(out.values()), dropped


def existing_tickers() -> tuple[set[str], str]:
    """Alle tickers die de DB al kent (ook inactief/delisted). Zonder
    DATABASE_URL: de actieve set uit de live API (alleen goed genoeg voor dry-run)."""
    if os.environ.get("DATABASE_URL"):
        from engine import db
        with db.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT ticker FROM stocks")
            return {r[0] for r in cur.fetchall()}, "database"
    rows = _http_get(f"{APP_URL}/api/dashboard").json()
    return {r["ticker"] for r in rows}, "live API (alleen actieve tickers!)"


def probe_batch(records: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Laat de server testen welke kandidaat-symbolen Yahoo kent.

    Dit is de vertaalslag van beurssymbool naar Yahoo-symbool: de beurs noemt
    Balder "BALD B", Yahoo noemt het FAST-B.ST. Zonder deze stap importeren we
    symbolen die nooit data opleveren en het dashboard blijven vervuilen.

    Returns (te importeren records, onvindbare records).
    """
    verify = os.environ.get("IMPORT_TLS_VERIFY", "1") != "0"
    by_ticker = {r["ticker"]: r for r in records}
    resolved: set[str] = set()
    unknown: set[str] = set()  # chunk mislukt: geen oordeel, dus niet importeren

    tickers = list(by_ticker)
    for start in range(0, len(tickers), 200):
        chunk = tickers[start:start + 200]
        r = requests.post(f"{APP_URL}/api/stocks/probe", json={"tickers": chunk},
                          headers=UA, timeout=180, verify=verify)
        if r.status_code != 200:
            print(f"  [WAARSCHUWING] probe mislukt ({r.status_code}); "
                  f"{len(chunk)} tickers overgeslagen", file=sys.stderr)
            unknown.update(chunk)
            continue
        data = r.json()
        resolved.update(data.get("resolved") or [])
        if data.get("chunks_failed"):
            print(f"  [WAARSCHUWING] {data['chunks_failed']} deelchunk(s) faalden bij Yahoo; "
                  "die tickers blijven liggen voor een volgende run", file=sys.stderr)

    keep = [by_ticker[t] for t in tickers if t in resolved]
    lost = [by_ticker[t] for t in tickers if t not in resolved and t not in unknown]
    pct = (100 * len(keep) // len(tickers)) if tickers else 0
    print(f"\nProbe: {len(keep)}/{len(tickers)} symbolen leveren een koers op ({pct}%)")
    if lost:
        print(f"  {len(lost)} onvindbaar bij Yahoo, o.a.: "
              + ", ".join(r["ticker"] for r in lost[:12]))
    return keep, lost


# ---------------------------------------------------------------- main

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Universe-import vanaf officiële beurslijsten (fase 4).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Zie de module-docstring voor bronnen en werkwijze.")
    ap.add_argument("--source", required=True,
                    choices=["euronext", "xetra", "nasdaq-nordic", "baltic", "gpw"])
    ap.add_argument("--file", type=Path, default=None,
                    help="Lokaal bronbestand i.p.v. auto-download")
    ap.add_argument("--limit", type=int, default=250,
                    help="Max nieuwe tickers per run (default 250, fase 4-tempo)")
    ap.add_argument("--dry-run", action="store_true", help="Alleen rapporteren, niets schrijven")
    ap.add_argument("--include-growth", action="store_true",
                    help="Ook Growth/Access/Scale/First North Baltic-segmenten")
    ap.add_argument("--out", type=Path, default=None,
                    help="Schrijf de nieuwe tickers (vóór --limit) als CSV")
    ap.add_argument("--probe", action="store_true",
                    help="Test kandidaten eerst via /api/stocks/probe op de server "
                         "en importeer alleen de tickers waarvoor Yahoo een koers "
                         "kent (sterk aanbevolen; werkt ook samen met --dry-run)")
    args = ap.parse_args()

    if args.file and not args.file.exists():
        print(f"FOUT: bestand niet gevonden: {args.file}", file=sys.stderr)
        return 1

    print(f"Bron: {args.source}" + (f" ({args.file})" if args.file else " (auto-download)"))
    if args.source == "euronext":
        records = fetch_euronext(args.file)
        if not args.include_growth:
            records = [r for r in records if r["segment"] == "regulated"]
    elif args.source == "xetra":
        records = fetch_xetra(args.file, args.include_growth)
    elif args.source == "nasdaq-nordic":
        records = fetch_nasdaq_nordic(args.file)
    elif args.source == "baltic":
        records = fetch_baltic(args.file, args.include_growth)
    else:
        records = fetch_gpw(args.file)

    print(f"{len(records)} geldige noteringen in de bron")
    records, isin_dropped = dedupe_isin(records)
    if isin_dropped:
        print(f"{len(isin_dropped)} secundaire noteringen gedropt (zelfde ISIN): "
              + ", ".join(isin_dropped[:10]) + ("..." if len(isin_dropped) > 10 else ""))

    from engine.remap_rules import REMAP_RULES
    known_secondary = [r["ticker"] for r in records if r["ticker"] in REMAP_RULES]
    if known_secondary:
        records = [r for r in records if r["ticker"] not in REMAP_RULES]
        print(f"{len(known_secondary)} bekende secondary listings overgeslagen (remap_rules): "
              + ", ".join(known_secondary))

    existing, existing_src = existing_tickers()
    print(f"{len(existing)} bestaande tickers geladen uit {existing_src}")
    new = [r for r in records if r["ticker"] not in existing]
    already = len(records) - len(new)

    per_market: dict[str, int] = {}
    for r in new:
        per_market[r["market"]] = per_market.get(r["market"], 0) + 1
    print(f"\nAl bekend: {already} · NIEUW: {len(new)}")
    for m, n in sorted(per_market.items(), key=lambda x: -x[1]):
        print(f"  {m:<4} {n:>5}")

    if args.out:
        with args.out.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["ticker", "name", "isin", "market", "currency", "sector", "segment"])
            w.writeheader()
            w.writerows(new)
        print(f"Nieuwe tickers weggeschreven naar {args.out}")

    batch = new[:args.limit]
    if len(new) > args.limit:
        print(f"--limit {args.limit}: {len(new) - args.limit} blijven liggen voor een volgende run")

    if args.probe:
        batch, unresolved = probe_batch(batch)
        if unresolved and args.out:
            reject_path = args.out.with_name(args.out.stem + "_onvindbaar.csv")
            with reject_path.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=["ticker", "name", "isin", "market", "currency", "sector", "segment"])
                w.writeheader()
                w.writerows(unresolved)
            print(f"Onvindbare symbolen weggeschreven naar {reject_path} "
                  "(kandidaten voor handmatige remap)")

    if args.dry_run:
        print(f"\n[DRY RUN] Zou {len(batch)} tickers toevoegen. Voorbeeld:")
        for r in batch[:10]:
            print(f"  {r['ticker']:<12} {r['name'] or '':<40} {r['isin'] or ''}")
        return 0

    if not os.environ.get("DATABASE_URL"):
        print("FOUT: echt importeren vereist DATABASE_URL.", file=sys.stderr)
        return 1

    from engine.db import init_db, upsert_stock
    init_db()
    today_iso = date.today().isoformat()
    ok, errors = 0, []
    for r in batch:
        try:
            upsert_stock(
                r["ticker"], name=r["name"], market=r["market"],
                currency=r["currency"], financial_currency=r["currency"],
                isin=r["isin"], sector=r["sector"],
                active=1, added_date=today_iso,
            )
            ok += 1
        except Exception as e:  # noqa: BLE001 — doorgaan met de rest
            errors.append((r["ticker"], str(e)))
    print(f"\n{ok} tickers toegevoegd; de nachtelijke fundamentals-rotatie pakt ze met voorrang op.")
    if errors:
        print(f"{len(errors)} fouten:")
        for tk, msg in errors[:20]:
            print(f"  {tk}: {msg}")
    return 0 if not errors else 2


if __name__ == "__main__":
    sys.exit(main())
